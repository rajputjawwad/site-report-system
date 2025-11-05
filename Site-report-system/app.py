from flask import Flask, render_template, request, send_file
import gspread
from google.oauth2.service_account import Credentials
from io import BytesIO
from zipfile import ZipFile
import openpyxl
import win32com.client as win32
import os

app = Flask(__name__)

# ===== GOOGLE SHEET CONNECTION =====
SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets.readonly",
    "https://www.googleapis.com/auth/drive.readonly"
]
creds = Credentials.from_service_account_file("credentials.json", scopes=SCOPES)
client = gspread.authorize(creds)
sheet = client.open("Site_database").sheet1  # Sheet name must match exactly


@app.route('/')
def index():
    sites = sheet.get_all_records()
    return render_template('index.html', sites=sites)


@app.route('/generate', methods=['POST'])
def generate():
    month_year = request.form['month_year']
    selected_sites = request.form.getlist('sites')

    all_sites = sheet.get_all_records()
    if "ALL" in selected_sites:
        sites_to_generate = all_sites
    else:
        sites_to_generate = [s for s in all_sites if s['site_name'] in selected_sites]

    if not sites_to_generate:
        return "❌ No site selected."

    buffer = BytesIO()

    with ZipFile(buffer, "w") as zipf:
        for site in sites_to_generate:
            site_name = site.get('site_name') or site.get('Site Name') or "Unknown_Site"
            site_address = next((v for k, v in site.items() if 'address' in k.lower()), " ")

            for i in range(1, 8):
                pdf_path = fill_excel_and_export(site_name, site_address, month_year, i)
                if pdf_path:
                    zipf.write(pdf_path, os.path.basename(pdf_path))

    buffer.seek(0)
    return send_file(
        buffer,
        as_attachment=True,
        download_name=f"Reports_{month_year}.zip",
        mimetype="application/zip"
    )


def fill_excel_and_export(site_name, site_address, month_year, report_no):
    """Fill Excel template cells and export to PDF."""
    template_path = f"Static/Report{report_no}.xlsx"
    if not os.path.exists(template_path):
        print(f"⚠️ Missing: {template_path}")
        return None

    output_xlsx = f"temp_{site_name}_Report{report_no}.xlsx"
    output_pdf = f"temp_{site_name}_Report{report_no}.pdf"

    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # ====== FILL CELLS BASED ON REPORT TYPE ======
    if 1 <= report_no <= 5:
        # Reports 1–5
        ws["I5"] = site_address      # Establishment
        ws["I8"] = site_address      # Principal Employer
        ws["I10"] = site_address     # Immediate Employer
        ws["D10"] = month_year       # Date
    elif report_no == 6:
        # Example placeholders — adjust later
        ws["D5"] = month_year
        ws["I5"] = site_address
    elif report_no == 7:
        ws["D6"] = month_year
        ws["I5"] = site_address

    wb.save(output_xlsx)

    # ===== EXPORT EXCEL TO PDF (Via Excel App) =====
    excel = win32.gencache.EnsureDispatch('Excel.Application')
    excel.Visible = False
    wb_obj = excel.Workbooks.Open(os.path.abspath(output_xlsx))
    wb_obj.ExportAsFixedFormat(0, os.path.abspath(output_pdf))
    wb_obj.Close(False)
    excel.Quit()

    os.remove(output_xlsx)  # optional cleanup
    return output_pdf


if __name__ == '__main__':
    app.run(debug=True)
