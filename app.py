from flask import Flask, render_template, request, send_file, jsonify
import gspread
from google.oauth2.service_account import Credentials
from io import BytesIO
from zipfile import ZipFile
import openpyxl
import pdfkit
import os

# Detect wkhtmltopdf automatically or use default path on Render
try:
    config = pdfkit.configuration(wkhtmltopdf="/usr/bin/wkhtmltopdf")
except:
    config = None

app = Flask(__name__)

# ===== GOOGLE SHEET CONNECTION =====
SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive"
]
import json, os
from google.oauth2.service_account import Credentials

service_account_info = json.loads(os.environ["GOOGLE_CREDENTIALS"])
creds = Credentials.from_service_account_info(service_account_info, scopes=SCOPES)
client = gspread.authorize(creds)
sheet = client.open("Site_database").sheet1

# ===== REPORT NAMES =====
REPORT_NAMES = {
    1: "REGISTER OF DEDUCTIONS FOR DAMAGE OR LOSS",
    2: "REGISTER OF ADVANCES",
    3: "REGISTER OF OVERTIME",
    4: "REGISTER OF FINES",
    5: "REGISTER OF ACCIDENTS AND DANGEROUS OCCURENCES",
    6: "ACCIDENT BOOK",
    7: "Maternity Benefit Register"
}

@app.route("/add_site", methods=["POST"])
def add_site():
    data = request.get_json()
    site_name = data.get("site_name")
    site_address = data.get("site_address")

    if not site_name or not site_address:
        return jsonify({"message": "Both fields required!"}), 400

    sheet.append_row([site_name, site_address])
    return jsonify({"message": "✅ Site added successfully!"}), 200

@app.route('/')
def index():
    sites = sheet.get_all_records()
    site_names = [s.get('site_name') or s.get('Site Name') or "Unknown Site" for s in sites]
    return render_template('index.html', sites=site_names)


@app.route('/generate', methods=['POST'])
def generate():
    month_year = request.form['month_year']
    selected_sites = request.form.getlist('sites')

    all_sites = sheet.get_all_records()

    if "ALL" in selected_sites:
        sites_to_generate = all_sites
    else:
        sites_to_generate = [s for s in all_sites if (s.get('site_name') or s.get('Site Name')) in selected_sites]

    if not sites_to_generate:
        return "❌ No site selected."

    buffer = BytesIO()
    with ZipFile(buffer, "w") as zipf:
        for site in sites_to_generate:
            site_name = site.get('site_name') or site.get('Site Name') or "Unknown_Site"
            site_address = next((v for k, v in site.items() if 'address' in k.lower()), " ")
            folder_name = site_name.replace(" ", "_")

            for i in range(1, 8):
                pdf_path = fill_excel_and_export(site_name, site_address, month_year, i)
                if pdf_path:
                    report_filename = f"{REPORT_NAMES[i]}.pdf"
                    zipf.write(pdf_path, os.path.join(folder_name, report_filename))
                    os.remove(pdf_path)

    buffer.seek(0)
    return send_file(
        buffer,
        as_attachment=True,
        download_name=f"Reports_{month_year}.zip",
        mimetype="application/zip"
    )


def fill_excel_and_export(site_name, site_address, month_year, report_no):
    try:
        template_path = f"static/Report{report_no}.xlsx"
        if not os.path.exists(template_path):
            print(f"⚠️ Missing template: {template_path}")
            return None

        wb = openpyxl.load_workbook(template_path)
        ws = wb.active

        if 1 <= report_no <= 5:
            ws["I5"] = site_address
            ws["I8"] = site_address
            ws["I10"] = site_address
            ws["D10"] = month_year
        elif report_no == 6:
            ws["A5"] = f"Name and address of the Establishment:- {site_address}"
            ws["A9"] = f"There are no Accident for the Month {month_year}"
        elif report_no == 7:
            ws["A4"] = f"Name and address of the Establishment:- {site_address}"
            ws["R4"] = month_year

        output_xlsx = f"temp_{site_name}_Report{report_no}.xlsx"
        wb.save(output_xlsx)

        # Convert Excel to HTML
        html_content = "<html><body><table border='1'>"
        for row in ws.iter_rows(values_only=True):
            html_content += "<tr>" + "".join(
                [f"<td>{str(cell) if cell else ''}</td>" for cell in row]
            ) + "</tr>"
        html_content += "</table></body></html>"

        html_path = f"temp_{site_name}_Report{report_no}.html"
        with open(html_path, "w") as f:
            f.write(html_content)

        output_pdf = f"temp_{site_name}_Report{report_no}.pdf"
        pdfkit.from_file(html_path, output_pdf, configuration=config)


        os.remove(html_path)
        os.remove(output_xlsx)
        return output_pdf

    except Exception as e:
        print(f"❌ Error generating Report {report_no} for {site_name}: {e}")
        return None


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000)
